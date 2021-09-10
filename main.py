import openpyxl
from openpyxl.styles import PatternFill

from dict import ts, dal_chadar
try:
    filename = "2020-06-14 - STOCK UPDATE.xlsx"

    wb = openpyxl.load_workbook(filename)
    print(wb.sheetnames)

    def pri_data_TSHIRT():
        """
        runs if user enter TSHIRT in sheet as tshirt have 2 value
        1. SKUID
        2. Size
        """
        def substract_value_tshirt():
            '''substract value of sold quantity from stock then saves'''
            subvalue_tshirt = int(input("Enter quantity of sold item: "))
            CHVsub_TSHIRT = data_TSHIRT - subvalue_tshirt
            sh1[ts[SKUID_TSHIRT][SIZE_TSHIRT]].value = CHVsub_TSHIRT
            if CHVsub_TSHIRT >= 5:
                sh1[ts[SKUID_TSHIRT][SIZE_TSHIRT]].fill = PatternFill(
                    "solid", fgColor="7EFF1F")
            elif CHVsub_TSHIRT >= 3:
                sh1[ts[SKUID_TSHIRT][SIZE_TSHIRT]].fill = PatternFill(
                    "solid", fgColor="ffff00")

            elif CHVsub_TSHIRT <= 2:
                sh1[ts[SKUID_TSHIRT][SIZE_TSHIRT]].fill = PatternFill(
                    "solid", fgColor="FF240A")
            wb.save(filename)

        def add_value_tshirt():
            '''Add's quantity of new stock that has arrived then saves '''
            addvalue_tshirt = int(input("Enter arrived quantity of item: "))
            CHVadd_TSHIRT = data_TSHIRT + addvalue_tshirt
            sh1[ts[SKUID_TSHIRT][SIZE_TSHIRT]].value = CHVadd_TSHIRT
            if CHVadd_TSHIRT >= 5:
                sh1[ts[SKUID_TSHIRT][SIZE_TSHIRT]].fill = PatternFill(
                    "solid", fgColor="7EFF1F")
            elif CHVadd_TSHIRT >= 3:
                sh1[ts[SKUID_TSHIRT][SIZE_TSHIRT]].fill = PatternFill(
                    "solid", fgColor="ffff00")
            elif CHVadd_TSHIRT <= 2:
                sh1[ts[SKUID_TSHIRT][SIZE_TSHIRT]].fill = PatternFill(
                    "solid", fgColor="FF240A")

            elif CHVadd_TSHIRT <= 2:
                sh1[ts[SKUID_TSHIRT][SIZE_TSHIRT]].fill = PatternFill(
                    "solid", fgColor="FF240A")
            wb.save(filename)

        SKUID_TSHIRT = input("Enter SKUID: ")
        SIZE_TSHIRT = input("Enter Size: ")
        data_TSHIRT = sh1[ts[SKUID_TSHIRT][SIZE_TSHIRT]].value
        print("The current avaliable stock is: ", data_TSHIRT)
        what_to_do_tshirt = input(
            """Enter "Stockup" if adding stock of enter "Selling" if packing order: """
        )

        if what_to_do_tshirt == "Selling":
            substract_value_tshirt()
        if what_to_do_tshirt == "Stockup":
            add_value_tshirt()

    def pri_data_DAL_CHADAR():
        """
        FUNCTION TO PRINT VALUE IF USER
            ENTERS DAL CHADAR AS SKUID
        """
        def substract_value_dalchadar():
            """
            substract value form a cell in dal chadar sheet
            """
            subvalue_dalchadar = int(input("Enter quantity of sold item: "))
            CHVsub_dalchadar = data_dalchadar - subvalue_dalchadar
            sh1[dal_chadar[SKUID_DAL_CHADAR]].value = CHVsub_dalchadar
            if CHVsub_dalchadar >= 5:
                sh1[dal_chadar[SKUID_DAL_CHADAR]].fill = PatternFill(
                    "solid", fgColor="7EFF1F")
            elif CHVsub_dalchadar >= 3:
                sh1[dal_chadar[SKUID_DAL_CHADAR]].fill = PatternFill(
                    "solid", fgColor="ffff00")

            elif CHVsub_dalchadar <= 2:
                sh1[dal_chadar[SKUID_DAL_CHADAR]].fill = PatternFill(
                    "solid", fgColor="FF240A")
            wb.save(filename)

        def add_value_dalchadar():
            addvalue_dalchadar = int(input("Enter arrived quantity of item: "))
            CHVadd_dalchadar = addvalue_dalchadar + data_dalchadar
            sh1[dal_chadar[SKUID_DAL_CHADAR]].value = CHVadd_dalchadar
            if CHVadd_dalchadar >= 5:
                sh1[dal_chadar[SKUID_DAL_CHADAR]].fill = PatternFill(
                    "solid", fgColor="7EFF1F")
            elif CHVadd_dalchadar >= 3:
                sh1[dal_chadar[SKUID_DAL_CHADAR]].fill = PatternFill(
                    "solid", fgColor="ffff00")

            elif CHVadd_dalchadar <= 2:
                sh1[dal_chadar[SKUID_DAL_CHADAR]].fill = PatternFill(
                    "solid", fgColor="FF240A")
            wb.save(filename)

        SKUID_DAL_CHADAR = input("Enter SKUID")
        data_dalchadar = sh1[dal_chadar[SKUID_DAL_CHADAR]].value
        print("The current avaliable stock is: ", data_dalchadar)
        what_to_do_dalchadar = input(
            """Enter "Stockup" if adding stock of enter "Selling" if packing order: """
        )
        if what_to_do_dalchadar == "Selling":
            substract_value_dalchadar()
        if what_to_do_dalchadar == "Stockup":
            add_value_dalchadar()

    #---------------------------------------------------------------
    User_sheet_input = input("Enter Sheet name: ")
    sh1 = wb[User_sheet_input]
    if User_sheet_input == "T-SHIRT":
        pri_data_TSHIRT()
    if User_sheet_input == "DAL CHADAR":
        pri_data_DAL_CHADAR()
except Exception as e:
    print(e)